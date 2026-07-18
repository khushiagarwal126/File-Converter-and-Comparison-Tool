"""
tests/unit/test_convert.py
───────────────────────────
Unit tests for POST /api/v1/convert.

Strategy
────────
- All tests use Flask's test client (no real HTTP server).
- Source files are created in a per-test tmp directory so nothing
  touches real uploads/temp or outputs/.
- Output files are also written to the same tmp directory.
- We test the full service path end-to-end: request → service → disk → response.
"""

import io
import json
import os
import shutil
import tempfile
import pytest

from app.main import create_app
from app.core.config import settings


# ─────────────────────────────────────────────────────────────────────────────
# Fixtures
# ─────────────────────────────────────────────────────────────────────────────

@pytest.fixture
def app(tmp_path):
    """Fresh app + isolated tmp dirs for each test."""
    flask_app = create_app()
    flask_app.config["TESTING"] = True
    settings.UPLOAD_DIR = str(tmp_path / "uploads")
    settings.OUTPUT_DIR = str(tmp_path / "outputs")
    os.makedirs(settings.UPLOAD_DIR, exist_ok=True)
    os.makedirs(settings.OUTPUT_DIR, exist_ok=True)
    yield flask_app
    # Restore defaults
    settings.UPLOAD_DIR = os.getenv("UPLOAD_DIR", "uploads/temp")
    settings.OUTPUT_DIR = os.getenv("OUTPUT_DIR", "outputs")


@pytest.fixture
def client(app):
    return app.test_client()


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

CSV_CONTENT  = b"name,age,city\nAlice,30,London\nBob,25,Paris\n"
JSON_CONTENT = b'[{"name":"Alice","age":30,"city":"London"},{"name":"Bob","age":25,"city":"Paris"}]'
XML_CONTENT  = (
    b"<?xml version='1.0' encoding='utf-8'?>\n"
    b"<data><row><name>Alice</name><age>30</age><city>London</city></row>"
    b"<row><name>Bob</name><age>25</age><city>Paris</city></row></data>"
)
TXT_CONTENT  = b"Hello World\nThis is line two\nLine three"


def write_source(filename: str, content: bytes) -> str:
    """Write a source file into UPLOAD_DIR and return its absolute path."""
    path = os.path.join(settings.UPLOAD_DIR, filename)
    with open(path, "wb") as f:
        f.write(content)
    return path


def post_convert(client, file_path: str, target_format: str):
    """POST to /api/v1/convert with JSON body."""
    return client.post(
        f"{settings.API_PREFIX}/convert",
        data=json.dumps({
            "file_path":     file_path,
            "target_format": target_format,
        }),
        content_type="application/json",
    )


# ─────────────────────────────────────────────────────────────────────────────
# Request validation
# ─────────────────────────────────────────────────────────────────────────────

class TestRequestValidation:

    def test_no_body_returns_400(self, client):
        r = client.post(f"{settings.API_PREFIX}/convert")
        assert r.status_code == 400

    def test_no_body_error_code(self, client):
        r = client.post(f"{settings.API_PREFIX}/convert")
        assert r.get_json()["error"] == "MISSING_BODY"

    def test_empty_json_object_file_path_missing(self, client):
        r = client.post(
            f"{settings.API_PREFIX}/convert",
            data=json.dumps({"target_format": "json"}),
            content_type="application/json",
        )
        assert r.status_code == 400
        assert r.get_json()["error"] == "MISSING_FILE_PATH"

    def test_missing_target_format(self, client):
        src = write_source("test.csv", CSV_CONTENT)
        r = client.post(
            f"{settings.API_PREFIX}/convert",
            data=json.dumps({"file_path": src}),
            content_type="application/json",
        )
        assert r.status_code == 400
        assert r.get_json()["error"] == "MISSING_TARGET_FORMAT"

    def test_nonexistent_file_returns_404(self, client):
        r = post_convert(client, "/tmp/does_not_exist.csv", "json")
        assert r.status_code == 404
        assert r.get_json()["error"] == "FILE_NOT_FOUND"

    def test_unsupported_target_format(self, client):
        src = write_source("test.csv", CSV_CONTENT)
        r = post_convert(client, src, "docx")
        assert r.status_code == 400
        assert r.get_json()["error"] == "UNSUPPORTED_TARGET_FORMAT"

    def test_same_format_returns_400(self, client):
        src = write_source("test.csv", CSV_CONTENT)
        r = post_convert(client, src, "csv")
        assert r.status_code == 400
        assert r.get_json()["error"] == "SAME_FORMAT"

    def test_get_method_not_allowed(self, client):
        r = client.get(f"{settings.API_PREFIX}/convert")
        assert r.status_code == 405

    def test_error_envelope_shape(self, client):
        r = client.post(f"{settings.API_PREFIX}/convert")
        d = r.get_json()
        assert d["success"] is False
        assert "error" in d
        assert "message" in d


# ─────────────────────────────────────────────────────────────────────────────
# Success response shape
# ─────────────────────────────────────────────────────────────────────────────

class TestSuccessResponseShape:

    def test_returns_200(self, client):
        src = write_source("data.csv", CSV_CONTENT)
        assert post_convert(client, src, "json").status_code == 200

    def test_success_true(self, client):
        src = write_source("data.csv", CSV_CONTENT)
        assert post_convert(client, src, "json").get_json()["success"] is True

    def test_message_field(self, client):
        src = write_source("data.csv", CSV_CONTENT)
        assert "converted" in post_convert(client, src, "json").get_json()["message"].lower()

    def test_data_has_conversion_id(self, client):
        src = write_source("data.csv", CSV_CONTENT)
        d = post_convert(client, src, "json").get_json()["data"]
        assert len(d["conversion_id"]) == 36   # UUID4

    def test_data_has_source_format(self, client):
        src = write_source("data.csv", CSV_CONTENT)
        d = post_convert(client, src, "json").get_json()["data"]
        assert d["source_format"] == "csv"

    def test_data_has_target_format(self, client):
        src = write_source("data.csv", CSV_CONTENT)
        d = post_convert(client, src, "json").get_json()["data"]
        assert d["target_format"] == "json"

    def test_data_has_output_path(self, client):
        src = write_source("data.csv", CSV_CONTENT)
        d = post_convert(client, src, "json").get_json()["data"]
        assert "output_path" in d

    def test_data_has_size_bytes(self, client):
        src = write_source("data.csv", CSV_CONTENT)
        d = post_convert(client, src, "json").get_json()["data"]
        assert d["size_bytes"] > 0

    def test_data_has_output_filename(self, client):
        src = write_source("data.csv", CSV_CONTENT)
        d = post_convert(client, src, "json").get_json()["data"]
        assert d["output_filename"].endswith(".json")

    def test_output_file_exists_on_disk(self, client):
        src = write_source("data.csv", CSV_CONTENT)
        d = post_convert(client, src, "json").get_json()["data"]
        assert os.path.isfile(d["output_path"])

    def test_two_conversions_have_unique_ids(self, client):
        src1 = write_source("a.csv", CSV_CONTENT)
        src2 = write_source("b.csv", CSV_CONTENT)
        id1 = post_convert(client, src1, "json").get_json()["data"]["conversion_id"]
        id2 = post_convert(client, src2, "json").get_json()["data"]["conversion_id"]
        assert id1 != id2

    def test_output_filename_prefixed_with_conversion_id(self, client):
        src = write_source("data.csv", CSV_CONTENT)
        d = post_convert(client, src, "json").get_json()["data"]
        assert d["output_filename"].startswith(d["conversion_id"])


# ─────────────────────────────────────────────────────────────────────────────
# CSV as source
# ─────────────────────────────────────────────────────────────────────────────

class TestCsvSource:

    def _d(self, client, target):
        src = write_source("report.csv", CSV_CONTENT)
        return post_convert(client, src, target).get_json()["data"]

    def test_csv_to_json_is_valid_json(self, client):
        d = self._d(client, "json")
        content = open(d["output_path"]).read()
        parsed = json.loads(content)
        assert isinstance(parsed, list)
        assert len(parsed) == 2

    def test_csv_to_json_preserves_rows(self, client):
        d = self._d(client, "json")
        rows = json.loads(open(d["output_path"]).read())
        names = {r["name"] for r in rows}
        assert names == {"Alice", "Bob"}

    def test_csv_to_xml_contains_root_tag(self, client):
        d = self._d(client, "xml")
        content = open(d["output_path"]).read()
        assert "<data>" in content or "<root>" in content.lower() or "<row>" in content

    def test_csv_to_xml_contains_data(self, client):
        d = self._d(client, "xml")
        content = open(d["output_path"]).read()
        assert "Alice" in content
        assert "Bob" in content

    def test_csv_to_xlsx_is_valid_excel(self, client):
        import openpyxl
        d = self._d(client, "xlsx")
        wb = openpyxl.load_workbook(d["output_path"])
        ws = wb.active
        assert ws.max_row >= 2   # header + at least 1 data row

    def test_csv_to_xlsx_preserves_columns(self, client):
        import openpyxl
        d = self._d(client, "xlsx")
        wb = openpyxl.load_workbook(d["output_path"])
        headers = [c.value for c in wb.active[1]]
        assert "name" in headers
        assert "age" in headers

    def test_csv_to_txt_contains_headers(self, client):
        d = self._d(client, "txt")
        content = open(d["output_path"]).read()
        assert "name" in content
        assert "Alice" in content

    def test_csv_to_pdf_creates_file(self, client):
        d = self._d(client, "pdf")
        assert os.path.isfile(d["output_path"])
        assert d["size_bytes"] > 0

    def test_csv_to_pdf_is_valid_pdf(self, client):
        d = self._d(client, "pdf")
        header = open(d["output_path"], "rb").read(5)
        assert header == b"%PDF-"


# ─────────────────────────────────────────────────────────────────────────────
# JSON as source
# ─────────────────────────────────────────────────────────────────────────────

class TestJsonSource:

    def _d(self, client, target):
        src = write_source("data.json", JSON_CONTENT)
        return post_convert(client, src, target).get_json()["data"]

    def test_json_to_csv_has_header(self, client):
        d = self._d(client, "csv")
        content = open(d["output_path"]).read()
        assert "name" in content.splitlines()[0]

    def test_json_to_csv_preserves_rows(self, client):
        d = self._d(client, "csv")
        lines = [l for l in open(d["output_path"]).read().splitlines() if l]
        assert len(lines) == 3   # header + 2 data rows

    def test_json_to_xml_is_valid_xml(self, client):
        import xml.etree.ElementTree as ET
        d = self._d(client, "xml")
        ET.parse(d["output_path"])   # raises if invalid

    def test_json_to_xlsx(self, client):
        import openpyxl
        d = self._d(client, "xlsx")
        wb = openpyxl.load_workbook(d["output_path"])
        assert wb.active.max_row >= 2

    def test_json_to_txt(self, client):
        d = self._d(client, "txt")
        content = open(d["output_path"]).read()
        assert "Alice" in content

    def test_json_to_pdf(self, client):
        d = self._d(client, "pdf")
        assert open(d["output_path"], "rb").read(5) == b"%PDF-"


# ─────────────────────────────────────────────────────────────────────────────
# XML as source
# ─────────────────────────────────────────────────────────────────────────────

class TestXmlSource:

    def _d(self, client, target):
        src = write_source("data.xml", XML_CONTENT)
        return post_convert(client, src, target).get_json()["data"]

    def test_xml_to_csv_has_data(self, client):
        d = self._d(client, "csv")
        content = open(d["output_path"]).read()
        assert "Alice" in content

    def test_xml_to_json_is_valid(self, client):
        d = self._d(client, "json")
        rows = json.loads(open(d["output_path"]).read())
        assert len(rows) == 2

    def test_xml_to_xlsx(self, client):
        import openpyxl
        d = self._d(client, "xlsx")
        assert openpyxl.load_workbook(d["output_path"]).active.max_row >= 2

    def test_xml_to_txt(self, client):
        d = self._d(client, "txt")
        assert "Alice" in open(d["output_path"]).read()

    def test_xml_to_pdf(self, client):
        d = self._d(client, "pdf")
        assert open(d["output_path"], "rb").read(5) == b"%PDF-"


# ─────────────────────────────────────────────────────────────────────────────
# XLSX as source
# ─────────────────────────────────────────────────────────────────────────────

class TestXlsxSource:

    def _src(self):
        """Create a real xlsx file from CSV_CONTENT and return its path."""
        import pandas as pd
        path = os.path.join(settings.UPLOAD_DIR, "data.xlsx")
        pd.read_csv(io.StringIO(CSV_CONTENT.decode())).to_excel(
            path, index=False, engine="openpyxl"
        )
        return path

    def test_xlsx_to_csv(self, client):
        src = self._src()
        d = post_convert(client, src, "csv").get_json()["data"]
        assert "Alice" in open(d["output_path"]).read()

    def test_xlsx_to_json(self, client):
        src = self._src()
        d = post_convert(client, src, "json").get_json()["data"]
        rows = json.loads(open(d["output_path"]).read())
        assert len(rows) == 2

    def test_xlsx_to_xml(self, client):
        import xml.etree.ElementTree as ET
        src = self._src()
        d = post_convert(client, src, "xml").get_json()["data"]
        ET.parse(d["output_path"])

    def test_xlsx_to_txt(self, client):
        src = self._src()
        d = post_convert(client, src, "txt").get_json()["data"]
        assert "Alice" in open(d["output_path"]).read()

    def test_xlsx_to_pdf(self, client):
        src = self._src()
        d = post_convert(client, src, "pdf").get_json()["data"]
        assert open(d["output_path"], "rb").read(5) == b"%PDF-"


# ─────────────────────────────────────────────────────────────────────────────
# TXT as source
# ─────────────────────────────────────────────────────────────────────────────

class TestTxtSource:

    def _d(self, client, target):
        src = write_source("notes.txt", TXT_CONTENT)
        return post_convert(client, src, target).get_json()["data"]

    def test_txt_to_json(self, client):
        d = self._d(client, "json")
        rows = json.loads(open(d["output_path"]).read())
        assert len(rows) == 3
        assert "line" in rows[0]

    def test_txt_to_csv(self, client):
        d = self._d(client, "csv")
        content = open(d["output_path"]).read()
        assert "line" in content   # column header
        assert "Hello World" in content

    def test_txt_to_xml(self, client):
        import xml.etree.ElementTree as ET
        d = self._d(client, "xml")
        ET.parse(d["output_path"])

    def test_txt_to_xlsx(self, client):
        import openpyxl
        d = self._d(client, "xlsx")
        assert openpyxl.load_workbook(d["output_path"]).active.max_row >= 2

    def test_txt_to_pdf(self, client):
        d = self._d(client, "pdf")
        assert open(d["output_path"], "rb").read(5) == b"%PDF-"


# ─────────────────────────────────────────────────────────────────────────────
# PDF as source
# ─────────────────────────────────────────────────────────────────────────────

class TestPdfSource:

    def _make_pdf_src(self) -> str:
        """Create a minimal real PDF using reportlab and return its path."""
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.pagesizes import A4
        path = os.path.join(settings.UPLOAD_DIR, "document.pdf")
        c = rl_canvas.Canvas(path, pagesize=A4)
        c.drawString(72, 750, "name,age,city")
        c.drawString(72, 730, "Alice,30,London")
        c.drawString(72, 710, "Bob,25,Paris")
        c.save()
        return path

    def test_pdf_to_txt_creates_file(self, client):
        src = self._make_pdf_src()
        d = post_convert(client, src, "txt").get_json()["data"]
        assert os.path.isfile(d["output_path"])

    def test_pdf_to_txt_contains_text(self, client):
        src = self._make_pdf_src()
        d = post_convert(client, src, "txt").get_json()["data"]
        content = open(d["output_path"]).read()
        assert len(content) > 0

    def test_pdf_to_json_creates_file(self, client):
        src = self._make_pdf_src()
        r = post_convert(client, src, "json")
        assert r.status_code == 200
        assert os.path.isfile(r.get_json()["data"]["output_path"])

    def test_pdf_to_csv_creates_file(self, client):
        src = self._make_pdf_src()
        r = post_convert(client, src, "csv")
        assert r.status_code == 200

    def test_pdf_to_xml_creates_file(self, client):
        src = self._make_pdf_src()
        r = post_convert(client, src, "xml")
        assert r.status_code == 200

    def test_pdf_to_xlsx_creates_file(self, client):
        src = self._make_pdf_src()
        r = post_convert(client, src, "xlsx")
        assert r.status_code == 200


# ─────────────────────────────────────────────────────────────────────────────
# Output integrity
# ─────────────────────────────────────────────────────────────────────────────

class TestOutputIntegrity:

    def test_output_size_matches_disk(self, client):
        src = write_source("data.csv", CSV_CONTENT)
        d = post_convert(client, src, "json").get_json()["data"]
        assert d["size_bytes"] == os.path.getsize(d["output_path"])

    def test_output_in_configured_output_dir(self, client):
        src = write_source("data.csv", CSV_CONTENT)
        d = post_convert(client, src, "json").get_json()["data"]
        assert d["output_path"].startswith(os.path.abspath(settings.OUTPUT_DIR))

    def test_source_file_unchanged_after_conversion(self, client):
        src = write_source("data.csv", CSV_CONTENT)
        original_mtime = os.path.getmtime(src)
        post_convert(client, src, "json")
        assert os.path.getmtime(src) == original_mtime
        assert open(src, "rb").read() == CSV_CONTENT